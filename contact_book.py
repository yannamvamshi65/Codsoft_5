from openpyxl import Workbook,load_workbook
from time import sleep
wd=load_workbook("contact_book.xlsx")
ws=wd['Sheet1']
wb=wd.active
def add(ws,wb):
    name=input("Enter the name:\n")
    try:
        phno=int(input("Enter the Mobile number:\n"))
    except Exception:
        print("\"Mobile number only contains numbers\"")
        exit()
    email=input("Enter the Email id:\n")
    if email.endswith("@gmail.com"):
        print("Valid Email id")
    else:
        print("Invaild Email id")
        exit()
    Address=input("Enter the Address:\n")
    data=[name,phno,email,Address]
    for i in range(1,100):
        if ws['a'+str(i)].value==None:
            wb.append(data)
            wd.save("contact_book.xlsx")
            sleep(.5)
            print("\"DATA SAVED SUCCESSFULLY\"")
            exit()
def view(ws):
    rows=ws.iter_rows(min_row=2,min_col=1)
    for a,b,c,d in rows:
        print(a.value,b.value,c.value,d.value,sep="\t\t")
def search(ws):
    rows = ws.iter_rows(min_row=1, min_col=1)
    count = 0
    for a, b, c ,d in rows:
        count += 1
    name=input("enter the name:\n")
    for i in range(1,count+1):
        if ws['a'+str(i)].value==name:
            print("Name\t\tNumber\t\tEmail\t\tAddress")
            print(ws['a'+str(i)].value,":",ws['b'+str(i)].value,ws['c'+str(i)].value,ws['d'+str(i)].value,sep="\t")
            exit()
    else:
        print("\"NAME NOT FOUND\"")
        exit()
def update(ws,wb):
    name=input("Enter the Name you wish to Update:\n")
    for i in range(1,100):
        if ws['a'+str(i)].value==name:
            print("\nEnter Option\n1.Update Name\n2.Update Mobile no\n3.Update Email id\n4.Update Address")
            try:
                option=int(input("\nEnter the Option\n"))
            except Exception:
                print("\nOption Must be Number")
            if option==1:
                Name=input("Enter Name:\n")
                ws['a'+str(i)].value=Name
                wd.save("contact_book.xlsx")
                sleep(.5)
                print("Updated SUCESSFULLY")
                exit()
            elif option==2:
                try:
                    mobile=int(input("Enter the Number:\n"))
                except Exception:
                    print(Exception)
                ws['b'+str(i)].value=mobile
                wd.save("contact_book.xlsx")
                sleep(.5)
                print("Updated SUCESSFULLY")
                exit()
            elif option ==3:
                emil=input("Enter Email:\n")
                if emil.endswith("@gmail.com"):
                    print("Vaild Email id")
                    ws['c'+str(i)].value=emil
                    wd.save("contact_book.xlsx")
                    sleep(.5)
                    print("Updated SUCESSFULLY")
                    exit()
                else:
                    print("Invaild Email id")
            elif option==4:
                address=input("Enter the Address:\n")
                ws['b'+str(i)].value=address
                sleep(.5)
                print("Updated Successfully")
                exit()
    else:
        print("Name not found!!")

def delete(ws):
    name=input("Enter the Name:\n")
    if name=="all":
        ws.delete_rows(3,100)
        wd.save("contact_book.xlsx")
        sleep(.5)
        print("COMPLETE DATA DELETED")
        exit()
    else:
        for i in range(1,100):
            if ws['a'+str(i)].value==name:
                ws.delete_rows(idx=i,amount=1)
                wd.save("contact_book.xlsx")
                sleep(.5)
                print("DATA DELETED")
                exit()

print("*****YOUR CONTACT BOOK*****\n")
choice=int(input("SELECT THE SERVICE YOU WANT:\n1.ADD CONTACT\n2.CONTACT LIST\n3.SEARCH CONTACT\n4.UPDATE CONTACT\n5.DELETE CONTACT\n"))
match choice:
    case 1:
        add(ws,wb)
    case 2:
        sleep(.5)
        view(ws)
    case 3:
        search(ws)
    case 4:
        update(ws,wb)
    case 5:
        delete(ws)
    case _:
        print("Invalid input")